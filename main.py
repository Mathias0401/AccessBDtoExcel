import pyodbc,openpyxl, os, traceback,time
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

def main():

    pathBD = input("Ingrese la ruta a la Base de Datos Access: ")

    connStr = (r"DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};"fr"DBQ={pathBD};")
    cnxn = pyodbc.connect(connStr)

    query = "SELECT * FROM Tickets"
    dataf = pd.read_sql(query, cnxn)
    cnxn.close()

    try:
        if(os.path.isfile("./Stock.xlsx")):
            workbook = openpyxl.load_workbook('Stock.xlsx')
        else:
            workbook = openpyxl.Workbook()
        if("datos" not in workbook.sheetnames):
            workbook.create_sheet('datos')
            worksheet = workbook["datos"]
        else:
            del workbook['datos']
            workbook.create_sheet('datos')
            worksheet = workbook['datos']
            worksheet.delete_rows(1, worksheet.max_row)
    except:
        traceback.print_exc()
        input("Enter pa seguir...")

    try:
        os.system("clear||cls")
        fechaIni = input("Ingrese la fecha mínima a filtrar (Formato: 10-03-2024): ")
        fechaIni = fechaIni.split("-")

        fechaFin = input("\nIngrese la fecha máxima a filtrar (Formato: 10-03-2024): ")
        fechaFin = fechaFin.split("-")

        os.system("clear||cls")

        maxRow = 0
        if(len(dataf) != 0):
            filtered_df = dataf.loc[(dataf['Fecha'] >= f'{fechaIni[2]}-{fechaIni[1]}-{fechaIni[0]}') & (dataf['Fecha'] <= f'{fechaFin[2]}-{fechaFin[1]}-{fechaFin[0]}')]
            for r in dataframe_to_rows(filtered_df, header=True, index=False):
                maxRow+=1
                print(f"Trabajando... {maxRow}/{len(filtered_df)}")
                worksheet.append(r)
            if(maxRow>1):
                tab = Table(displayName="Table1", ref=f"A1:R{maxRow}")
                # Añado un estilo por defecto con filas a rayas
                style = TableStyleInfo(name="TableStyleMedium1", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                tab.tableStyleInfo = style
                worksheet.add_table(tab)
        else:
            print("No hay datos en la tabla")
        for row in range(2, worksheet.max_row+1):
            worksheet["{}{}".format("F", row)].number_format = 'General'
    except:
        traceback.print_exc()

    workbook.save('Stock.xlsx')
    input("Enter pa seguir...")

opt = -1
while(opt != "0"):
    if(opt == "1"):
        os.system("clear||cls")
        main()
    os.system("clear||cls")
    opt = input("1 - Consultar\n0 - Salir\n\nIngrese opción: ")